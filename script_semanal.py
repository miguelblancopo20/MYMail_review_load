import argparse
import json
import hashlib
import math
import shutil
import sys
import unicodedata
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill


def parse_args():
    parser = argparse.ArgumentParser(
        description="Scripts semanales + analisis Laura (validaciones/fichas/ejecuciones/subtematicas)"
    )

    subparsers = parser.add_subparsers(dest="command")

    semanal = subparsers.add_parser("semanal", help="Genera Excels semanales")
    semanal.add_argument("--fecha", default="4enero", help="Subcarpeta dentro de data/, por ejemplo: 4enero")
    semanal.add_argument(
        "--no-laura",
        action="store_true",
        help="No ejecutar el analisis de Laura al correr el modo semanal",
    )
    semanal.add_argument(
        "--revision-size",
        type=int,
        default=150,
        help="Tamano del muestreo de revision (solo NO VALIDADOS)",
    )
    semanal.add_argument(
        "--no-revision",
        action="store_true",
        help="No generar validaciones_revision.xlsx al correr el modo semanal",
    )
    semanal.add_argument(
        "--revision-min-full",
        type=int,
        default=5,
        help="Si una combinacion (Automatismo, Segmento) tiene <= N casos NO VALIDADOS, se incluyen todos para evitar sesgo",
    )
    semanal.add_argument(
        "--no-stats",
        action="store_true",
        help="No generar el fichero basado en Plantilla_Stats.xlsx",
    )
    semanal.add_argument(
        "--stats-template",
        default=None,
        help="Ruta a Plantilla_Stats.xlsx (por defecto: ./Plantilla_Stats.xlsx)",
    )
    semanal.add_argument(
        "--stats-output",
        default=None,
        help="Nombre o ruta del fichero stats (por defecto: data/<fecha>/output/Stats_<fecha>.xlsx)",
    )

    stats = subparsers.add_parser("stats", help="Genera Stats_<fecha>.xlsx desde los Excels ya generados")
    stats.add_argument("--fecha", default="4enero", help="Subcarpeta dentro de data/, por ejemplo: 4enero")
    stats.add_argument(
        "--stats-template",
        default=None,
        help="Ruta a Plantilla_Stats.xlsx (por defecto: ./Plantilla_Stats.xlsx)",
    )
    stats.add_argument(
        "--stats-output",
        default=None,
        help="Nombre o ruta del fichero stats (por defecto: data/<fecha>/output/Stats_<fecha>.xlsx)",
    )

    laura = subparsers.add_parser("laura", help="Analisis IA/RPA y segmentacion (Laura)")
    laura.add_argument("--folder", default="data/4enero", help="Carpeta de trabajo dentro de data/")
    laura.add_argument(
        "--segmento-csv",
        default="data/CIF-Segmento_3.csv",
        help="CSV de CIF->SEGMENTO (debe tener NUMERO DOCUMENTO y SEGMENTO)",
    )
    laura.add_argument(
        "--output",
        default=None,
        help="Nombre o ruta del Excel de salida (por defecto: <folder>/output/mails_por_subtematica.xlsx)",
    )

    argv = sys.argv[1:]
    if argv and argv[0] in {"semanal", "stats", "laura"}:
        return parser.parse_args(argv)

    # Default: modo semanal
    return parser.parse_args(["semanal", *argv])


def get_dirs(fecha: str):
    base_dir = Path("data") / fecha
    output_dir = base_dir / "output"
    output_dir.mkdir(parents=True, exist_ok=True)
    return base_dir, output_dir


def fichas_levantadas(fecha: str):
    base_dir, output_dir = get_dirs(fecha)
    df = pd.read_csv(base_dir / "fichas_levantadas.csv")

    df["IdCorreo"] = df["IdCorreo"].apply(lambda x: x.split("-")[0])
    df = df.drop_duplicates(subset=["IdCorreo", "Automatismo"], keep="first")

    output_path = output_dir / "fichas_levantadas.xlsx"

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Data", index=False)

        if "Segmento" in df.columns:
            df_pivot_src = df.copy()
            df_pivot_src["Segmento"] = (
                df_pivot_src["Segmento"].fillna("(blank)").astype(str).replace({"": "(blank)"})
            )
            df_pivot_src["Automatismo"] = (
                df_pivot_src["Automatismo"].fillna("(blank)").astype(str).replace({"": "(blank)"})
            )
            df_pivot_src["IdCorreo"] = df_pivot_src["IdCorreo"].fillna("(blank)").astype(str).replace({"": "(blank)"})

            pivot = df_pivot_src.pivot_table(
                index="Automatismo",
                columns="Segmento",
                values="IdCorreo",
                aggfunc="count",
                fill_value=0,
                dropna=False,
            )
            pivot = _reordenar_pivot_blank(pivot, blank_label="(blank)", after_column="PE")

            pivot["Total"] = pivot.sum(axis=1)
            pivot.loc["Total"] = pivot.sum(axis=0)
            pivot = _reordenar_pivot_blank(pivot, blank_label="(blank)", after_column="PE", keep_total_last=True)
            pivot = pivot.reset_index()
            pivot.to_excel(writer, sheet_name="Pivot", index=False)
        else:
            pd.DataFrame(
                [{"WARN": "No existe la columna 'Segmento' en fichas_levantadas.csv; no se genera Pivot."}]
            ).to_excel(writer, sheet_name="Pivot", index=False)


def validados(fecha: str):
    base_dir, output_dir = get_dirs(fecha)
    df = pd.read_csv(base_dir / "validaciones.csv")
    df = df.drop_duplicates(subset=["IdCorreo"], keep="first")

    df_ia = pd.read_csv(base_dir / "ia-transacciones.csv", sep=";")
    df_ia = df_ia[["idLotus", "Location", "Sublocation", "Subject", "Question", "MailToAgent"]]

    df = pd.merge(df, df_ia, how="left", left_on="IdCorreo", right_on="idLotus")
    df["Faltan datos?"] = df.apply(lambda x: faltan_datos(x.Automatismo, x.MailToAgent), axis=1)

    df.sort_values(by=["Automatismo", "@timestamp"], ascending=[True, False], inplace=True)
    output_path = output_dir / "validaciones.xlsx"
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Data", index=False)

        df_pivot_src = df.copy()
        if "Segmento" in df_pivot_src.columns:
            df_pivot_src["Segmento"] = df_pivot_src["Segmento"].fillna("Sin Segmento")
        if "Validado" in df_pivot_src.columns:
            df_pivot_src["Validado"] = df_pivot_src["Validado"].fillna("Sin Validado")

        if all(col in df_pivot_src.columns for col in ["Automatismo", "Validado", "Segmento", "IdCorreo"]):
            pivot = df_pivot_src.pivot_table(
                index="Automatismo",
                columns=["Validado", "Segmento"],
                values="IdCorreo",
                aggfunc="count",
                fill_value=0,
            )
            pivot.columns = [f"{validado} | {segmento}" for validado, segmento in pivot.columns.to_list()]
            pivot = pivot.reset_index()
            pivot.to_excel(writer, sheet_name="Pivot", index=False)
        else:
            pd.DataFrame(
                [
                    {
                        "WARN": "Faltan columnas para generar Pivot (Automatismo, Validado, Segmento, IdCorreo).",
                    }
                ]
            ).to_excel(writer, sheet_name="Pivot", index=False)

    aplicar_estilos_validaciones_excel(output_path)


def faltan_datos(automatismo, mta):
    if "Apertura" not in str(automatismo):
        return False

    try:
        contexto = json.loads(mta)
        parametros = contexto["Parametros"]
        return not all([parametros.get("PERSONA CONTACTO"), parametros.get("TELEFONO CONTACTO")])
    except Exception:
        return False


def ejecuciones(fecha: str):
    base_dir, output_dir = get_dirs(fecha)
    df_validaciones = pd.read_excel(output_dir / "validaciones.xlsx")
    df_validaciones = df_validaciones.drop_duplicates(subset=["IdCorreo"], keep="first")
    df_validaciones = df_validaciones[
        ["IdCorreo", "Validado", "Documento", "Automatismo", "Segmento", "MatriculaAsesor"]
    ]
    df_validaciones = df_validaciones[df_validaciones["Validado"] == "VALIDADO"]

    automatismos = sorted(df_validaciones["Automatismo"].dropna().unique().tolist())

    df_orquestador_contexto = pd.read_csv(base_dir / "orquestador_contexto.csv")
    df_orquestador_contexto["IDU"] = df_orquestador_contexto["IDU"].apply(lambda x: str(x).split("-")[0])
    df_orquestador_contexto = df_orquestador_contexto.drop_duplicates(subset=["IDU"], keep="first")[["IDU"]]

    cruces = []

    for automatismo in automatismos:
        df_auto = df_validaciones[df_validaciones["Automatismo"] == automatismo]
        if df_auto.empty:
            continue

        df_cruce = pd.merge(df_auto, df_orquestador_contexto, how="left", left_on="IdCorreo", right_on="IDU")
        df_cruce = df_cruce.drop_duplicates(subset=["IdCorreo"], keep="first")
        df_cruce = df_cruce.rename(columns={"IdCorreo": "IdCorreo Validacion", "IDU": "IDU contexto"})
        cruces.append(df_cruce)

        total_validados = df_cruce["IdCorreo Validacion"].nunique()
        ejecutados = df_cruce["IDU contexto"].notna().sum()
        porcentaje = (100 * ejecutados / total_validados) if total_validados else 0

        print(f"**{automatismo}:")
        print(f"Validados: {total_validados}")
        print(f"Ejecutados: {ejecutados}")
        print(f"Porcentaje: {porcentaje}\n")

    output_path = output_dir / "ejecuciones.xlsx"

    if not cruces:
        with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
            pd.DataFrame().to_excel(writer, sheet_name="Data", index=False)
            pd.DataFrame([{"WARN": "No hay registros VALIDADO para generar ejecuciones/pivot."}]).to_excel(
                writer, sheet_name="Pivot", index=False
            )
        return

    df_out = pd.concat(cruces, ignore_index=True)
    df_out = df_out[
        ["IdCorreo Validacion", "Automatismo", "Segmento", "Documento", "MatriculaAsesor", "IDU contexto"]
    ]
    df_out["Encontrados"] = df_out["IDU contexto"].apply(lambda x: "SI" if pd.notna(x) else "NO")

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        df_out.to_excel(writer, sheet_name="Data", index=False)

        df_pivot_src = df_out.copy()
        df_pivot_src["Automatismo"] = df_pivot_src["Automatismo"].fillna("(blank)").astype(str).replace({"": "(blank)"})
        df_pivot_src["Segmento"] = df_pivot_src["Segmento"].fillna("(blank)").astype(str).replace({"": "(blank)"})
        df_pivot_src["Encontrados"] = df_pivot_src["Encontrados"].fillna("(blank)").astype(str).replace({"": "(blank)"})
        df_pivot_src["IdCorreo Validacion"] = (
            df_pivot_src["IdCorreo Validacion"].fillna("(blank)").astype(str).replace({"": "(blank)"})
        )

        pivot = df_pivot_src.pivot_table(
            index="Automatismo",
            columns=["Encontrados", "Segmento"],
            values="IdCorreo Validacion",
            aggfunc="count",
            fill_value=0,
            dropna=False,
        )
        pivot = _reordenar_pivot_blank_multiindex(pivot, blank_label="(blank)", after_segment="PE")

        pivot[("Total", "")] = pivot.sum(axis=1)
        pivot.loc["Total"] = pivot.sum(axis=0)

        pivot = _reordenar_pivot_blank_multiindex(pivot, blank_label="(blank)", after_segment="PE")

        pivot.columns = [
            "Total" if encontrados == "Total" else f"{encontrados} | {segmento}"
            for encontrados, segmento in pivot.columns.to_list()
        ]
        pivot = pivot.reset_index().rename(columns={"index": "Automatismo"})
        pivot.to_excel(writer, sheet_name="Pivot", index=False)


def aplicar_estilos_validaciones_excel(excel_path: Path):
    header_azul = {"idLotus", "Location", "Sublocation", "Subject", "Question", "MailToAgent", "Faltan datos?"}
    fill_azul = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    fill_verde = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")

    wb = load_workbook(excel_path)
    ws = wb["Data"] if "Data" in wb.sheetnames else wb.active

    for cell in ws[1]:
        if str(cell.value) in header_azul:
            cell.fill = fill_azul
        else:
            cell.fill = fill_verde

    wb.save(excel_path)


def _reordenar_pivot_blank(pivot: pd.DataFrame, blank_label: str, after_column: str, keep_total_last: bool = False):
    cols = list(pivot.columns)
    idx = list(pivot.index)

    total_label = "Total"
    cols_wo_special = [c for c in cols if c not in (blank_label, total_label)]
    idx_wo_special = [i for i in idx if i not in (blank_label, total_label)]

    ordered_cols = cols_wo_special
    if blank_label in cols:
        if after_column in ordered_cols:
            insert_at = ordered_cols.index(after_column) + 1
        else:
            insert_at = len(ordered_cols)
        ordered_cols = ordered_cols[:insert_at] + [blank_label] + ordered_cols[insert_at:]

    if keep_total_last and total_label in cols:
        ordered_cols = [c for c in ordered_cols if c != total_label] + [total_label]

    ordered_idx = idx_wo_special
    if blank_label in idx:
        ordered_idx = ordered_idx + [blank_label]
    if keep_total_last and total_label in idx:
        ordered_idx = ordered_idx + [total_label]

    return pivot.reindex(index=ordered_idx, columns=ordered_cols)


def _reordenar_pivot_blank_multiindex(
    pivot: pd.DataFrame,
    blank_label: str,
    after_segment: str,
    total_label: str = "Total",
):
    if not isinstance(pivot.columns, pd.MultiIndex) or pivot.columns.nlevels != 2:
        return pivot

    ordered_idx = [i for i in pivot.index if i not in (blank_label, total_label)]
    if blank_label in pivot.index:
        ordered_idx.append(blank_label)
    if total_label in pivot.index:
        ordered_idx.append(total_label)

    lvl0 = pivot.columns.get_level_values(0).unique().tolist()
    normal_lvl0 = [v for v in lvl0 if v != total_label]

    segments = pivot.columns.get_level_values(1).unique().tolist()
    segments_normal = [s for s in segments if s not in ("", blank_label)]
    if blank_label in segments:
        insert_at = segments_normal.index(after_segment) + 1 if after_segment in segments_normal else len(segments_normal)
        segments_order = segments_normal[:insert_at] + [blank_label] + segments_normal[insert_at:]
    else:
        segments_order = segments_normal

    ordered_cols = []
    for v0 in normal_lvl0:
        for seg in segments_order:
            col = (v0, seg)
            if col in pivot.columns:
                ordered_cols.append(col)

    for v0 in normal_lvl0:
        for seg in segments_normal:
            col = (v0, seg)
            if col in pivot.columns and col not in ordered_cols:
                ordered_cols.append(col)

    if total_label in lvl0:
        total_cols = [c for c in pivot.columns if c[0] == total_label]
        ordered_cols.extend([c for c in total_cols if c not in ordered_cols])

    return pivot.reindex(index=ordered_idx, columns=ordered_cols)


def _stable_int_seed(seed: int, *parts: str) -> int:
    h = hashlib.md5()
    h.update(str(seed).encode("utf-8"))
    for part in parts:
        h.update(b"|")
        h.update(str(part).encode("utf-8"))
    return int.from_bytes(h.digest()[:4], "big", signed=False)


def _load_or_init_revision_weights(weights_path: Path, automatismos: list[str], segmentos: list[str]):
    if weights_path.exists():
        with weights_path.open("r", encoding="utf-8") as f:
            weights = json.load(f)

        changed = False
        if "default" not in weights:
            weights["default"] = 1.0
            changed = True
        if "automatismo" not in weights or not isinstance(weights["automatismo"], dict):
            weights["automatismo"] = {}
            changed = True
        if "segmento" not in weights or not isinstance(weights["segmento"], dict):
            weights["segmento"] = {}
            changed = True
        if "pair" not in weights or not isinstance(weights["pair"], dict):
            weights["pair"] = {}
            changed = True

        for a in sorted(set(automatismos)):
            if a not in weights["automatismo"]:
                weights["automatismo"][a] = 1.0
                changed = True
        for s in sorted(set(segmentos)):
            if s not in weights["segmento"]:
                weights["segmento"][s] = 1.0
                changed = True

        if changed:
            with weights_path.open("w", encoding="utf-8") as f:
                json.dump(weights, f, ensure_ascii=False, indent=2)

        return weights

    weights = {
        "default": 1.0,
        "automatismo": {a: 1.0 for a in sorted(set(automatismos))},
        "segmento": {s: 1.0 for s in sorted(set(segmentos))},
        "pair": {},
    }
    with weights_path.open("w", encoding="utf-8") as f:
        json.dump(weights, f, ensure_ascii=False, indent=2)
    return weights


def _weight_for(weights: dict, automatismo: str, segmento: str) -> float:
    default = float(weights.get("default", 1.0))
    w_a = float(weights.get("automatismo", {}).get(automatismo, 1.0))
    w_s = float(weights.get("segmento", {}).get(segmento, 1.0))
    pair_key = f"{automatismo}||{segmento}"
    w_p = float(weights.get("pair", {}).get(pair_key, 1.0))
    w = default * w_a * w_s * w_p
    return max(w, 0.0)


def _allocate_samples(strata: list[dict], total_n: int) -> list[int]:
    total_n = max(int(total_n), 0)
    n_strata = len(strata)
    if n_strata == 0 or total_n == 0:
        return [0 for _ in strata]

    pops = [int(s["N"]) for s in strata]
    weights = [float(s["weight"]) for s in strata]
    eff = [p * w for p, w in zip(pops, weights)]
    eff_total = sum(eff)

    if eff_total <= 0:
        eff = pops[:]
        eff_total = sum(eff)
        if eff_total <= 0:
            return [0 for _ in strata]

    base = [0 for _ in strata]
    nonzero = [i for i, p in enumerate(pops) if p > 0]
    if total_n >= len(nonzero):
        for i in nonzero:
            base[i] = 1

    remaining = total_n - sum(base)
    if remaining <= 0:
        alloc = base[:]
    else:
        raw = [remaining * (e / eff_total) for e in eff]
        add = [int(math.floor(x)) for x in raw]
        alloc = [b + a for b, a in zip(base, add)]

        remainder = remaining - sum(add)
        frac = [r - math.floor(r) for r in raw]
        order = sorted(range(n_strata), key=lambda i: frac[i], reverse=True)
        for i in order:
            if remainder <= 0:
                break
            alloc[i] += 1
            remainder -= 1

    alloc = [min(a, p) for a, p in zip(alloc, pops)]
    deficit = total_n - sum(alloc)
    if deficit > 0:
        capacities = [p - a for p, a in zip(pops, alloc)]
        order = sorted(range(n_strata), key=lambda i: eff[i], reverse=True)
        for i in order:
            if deficit <= 0:
                break
            if capacities[i] <= 0:
                continue
            take = min(capacities[i], deficit)
            alloc[i] += take
            deficit -= take

    return alloc


def generar_validaciones_revision(fecha: str, sample_size: int = 150, min_full_stratum: int = 5):
    _, output_dir = get_dirs(fecha)
    validaciones_path = output_dir / "validaciones.xlsx"
    revision_path = output_dir / "validaciones_revision.xlsx"

    script_dir = Path(__file__).resolve().parent
    weights_path = script_dir / "validaciones_revision_pesos.json"
    legacy_weights_paths = [
        output_dir / "validaciones_revision_pesos.json",
        (Path("scriptsOld") / "validaciones_revision_pesos.json"),
    ]
    if not weights_path.exists():
        for legacy_path in legacy_weights_paths:
            if legacy_path.exists():
                weights_path.write_text(legacy_path.read_text(encoding="utf-8"), encoding="utf-8")
                break

    df = pd.read_excel(validaciones_path, sheet_name="Data")
    if "Validado" not in df.columns:
        with pd.ExcelWriter(revision_path, engine="openpyxl") as writer:
            pd.DataFrame([{"WARN": "No existe la columna 'Validado' en validaciones.xlsx"}]).to_excel(
                writer, sheet_name="Resumen", index=False
            )
        return

    df_no = df[df["Validado"].fillna("") != "VALIDADO"].copy()
    if df_no.empty:
        with pd.ExcelWriter(revision_path, engine="openpyxl") as writer:
            pd.DataFrame([{"WARN": "No hay NO VALIDADOS para muestrear."}]).to_excel(
                writer, sheet_name="Resumen", index=False
            )
        return

    if "Automatismo" not in df_no.columns:
        df_no["Automatismo"] = "(blank)"
    if "Segmento" not in df_no.columns:
        df_no["Segmento"] = "(blank)"
    if "IdCorreo" not in df_no.columns:
        df_no["IdCorreo"] = "(blank)"

    df_no["Automatismo"] = df_no["Automatismo"].fillna("(blank)").astype(str).replace({"": "(blank)"})
    df_no["Segmento"] = df_no["Segmento"].fillna("(blank)").astype(str).replace({"": "(blank)"})

    automatismos = sorted(df_no["Automatismo"].unique().tolist())
    segmentos = sorted(df_no["Segmento"].unique().tolist())
    weights = _load_or_init_revision_weights(weights_path, automatismos, segmentos)

    strata_df = (
        df_no.groupby(["Automatismo", "Segmento"], dropna=False)
        .size()
        .reset_index(name="N")
        .sort_values(["Automatismo", "Segmento"])
    )

    strata = []
    for row in strata_df.itertuples(index=False):
        w = _weight_for(weights, row.Automatismo, row.Segmento)
        strata.append({"Automatismo": row.Automatismo, "Segmento": row.Segmento, "N": int(row.N), "weight": w})

    target_n = min(int(sample_size), int(df_no.shape[0]))
    min_full_stratum = max(int(min_full_stratum), 0)

    small_strata = [s for s in strata if s["N"] > 0 and s["N"] <= min_full_stratum]
    small_total = sum(int(s["N"]) for s in small_strata)
    use_small_full = small_total > 0 and small_total <= target_n

    fixed_samples = []
    fixed_keys = set()
    if use_small_full:
        for s in small_strata:
            a = s["Automatismo"]
            seg = s["Segmento"]
            fixed_keys.add((a, seg))
            fixed_samples.append(df_no[(df_no["Automatismo"] == a) & (df_no["Segmento"] == seg)])

    remaining_strata = [s for s in strata if (s["Automatismo"], s["Segmento"]) not in fixed_keys]
    remaining_n = target_n - (small_total if use_small_full else 0)
    alloc = _allocate_samples(remaining_strata, remaining_n) if remaining_strata and remaining_n > 0 else []

    sampled_parts = []
    resumen_rows = []

    for s in strata:
        a = s["Automatismo"]
        seg = s["Segmento"]
        pop = s["N"]
        w = s["weight"]

        k = pop if (use_small_full and (a, seg) in fixed_keys) else 0
        resumen_rows.append(
            {
                "Automatismo": a,
                "Segmento": seg,
                "Poblacion_NO_VALIDADO": pop,
                "Peso": w,
                "Muestra": int(k),
            }
        )

    if use_small_full:
        sampled_parts.extend([df for df in fixed_samples if not df.empty])

    # actualizar Muestra para el resto usando alloc
    if remaining_strata and alloc:
        alloc_map = {(s["Automatismo"], s["Segmento"]): int(k) for s, k in zip(remaining_strata, alloc)}
        for row in resumen_rows:
            key = (row["Automatismo"], row["Segmento"])
            if key in alloc_map:
                row["Muestra"] = alloc_map[key]

        for s, k in zip(remaining_strata, alloc):
            if k <= 0:
                continue

            a = s["Automatismo"]
            seg = s["Segmento"]
            df_stratum = df_no[(df_no["Automatismo"] == a) & (df_no["Segmento"] == seg)]
            rs = _stable_int_seed(42, fecha, a, seg)
            sampled_parts.append(df_stratum.sample(n=int(k), random_state=rs))

    df_sample = pd.concat(sampled_parts, ignore_index=True) if sampled_parts else df_no.head(0)

    resumen_df = pd.DataFrame(resumen_rows)
    resumen_totales = pd.DataFrame(
        [
            {
                "Automatismo": "Total",
                "Segmento": "",
                "Poblacion_NO_VALIDADO": int(df_no.shape[0]),
                "Peso": "",
                "Muestra": int(df_sample.shape[0]),
            }
        ]
    )
    resumen_df = pd.concat([resumen_df, resumen_totales], ignore_index=True)

    with pd.ExcelWriter(revision_path, engine="openpyxl") as writer:
        df_sample.to_excel(writer, sheet_name="Data", index=False)
        resumen_df.to_excel(writer, sheet_name="Resumen", index=False)


def generar_stats_desde_outputs(fecha: str, template_path: str | None = None, output_path: str | None = None):
    base_dir, output_dir = get_dirs(fecha)

    if template_path:
        tpl_candidates = [Path(template_path)]
    else:
        script_dir = Path(__file__).resolve().parent
        tpl_candidates = [
            Path.cwd() / "Plantilla_Stats.xlsx",
            script_dir / "Plantilla_Stats.xlsx",
            script_dir.parent / "Plantilla_Stats.xlsx",
        ]

    tpl = next((p for p in tpl_candidates if p.exists()), None)

    if output_path:
        out = Path(output_path)
        if not out.is_absolute():
            out = (output_dir / out.name) if str(out.parent) == "." else (Path.cwd() / out).resolve()
    else:
        out = output_dir / f"Stats_{fecha}.xlsx"

    if tpl is None:
        checked = ", ".join(str(p) for p in tpl_candidates)
        print(f"[WARN] No existe la plantilla. Rutas comprobadas: {checked}")
        return

    out.parent.mkdir(parents=True, exist_ok=True)

    try:
        shutil.copy2(tpl, out)
    except PermissionError as exc:
        print(f"[WARN] No se puede leer/copiar la plantilla (¿esta abierta en Excel?): {tpl} ({exc})")
        return
    except FileNotFoundError as exc:
        if Path(tpl).exists():
            print(f"[WARN] No se pudo copiar la plantilla aunque existe (posible OneDrive/placeholder): {tpl} ({exc})")
        else:
            print(f"[WARN] No existe la plantilla: {tpl}")
        try:
            wb_tpl = load_workbook(tpl)
            wb_tpl.save(out)
        except Exception as exc2:
            print(f"[WARN] Fallback openpyxl fallo copiando plantilla: {exc2}")
            return
    except OSError as exc:
        print(f"[WARN] Error copiando plantilla: {tpl} ({exc})")
        try:
            wb_tpl = load_workbook(tpl)
            wb_tpl.save(out)
        except Exception as exc2:
            print(f"[WARN] Fallback openpyxl fallo copiando plantilla: {exc2}")
            return

    def safe_read_excel(path: Path, sheet: str):
        if not path.exists():
            return None
        try:
            return pd.read_excel(path, sheet_name=sheet)
        except Exception as exc:
            return pd.DataFrame([{"WARN": f"No se pudo leer {path.name}:{sheet} -> {exc}"}])

    validaciones_xlsx = output_dir / "validaciones.xlsx"
    fichas_xlsx = output_dir / "fichas_levantadas.xlsx"
    ejecuciones_xlsx = output_dir / "ejecuciones.xlsx"
    laura_xlsx = output_dir / "mails_por_subtematica.xlsx"
    revision_xlsx = output_dir / "validaciones_revision.xlsx"

    dfs: dict[str, pd.DataFrame] = {}

    df_val_data = safe_read_excel(validaciones_xlsx, "Data")
    if df_val_data is not None:
        resumen = {
            "fecha": fecha,
            "folder": str(base_dir),
            "validaciones_total": int(df_val_data.shape[0]),
            "validaciones_no_validados": int((df_val_data.get("Validado") != "VALIDADO").sum())
            if "Validado" in df_val_data.columns
            else "",
        }
        dfs["Resumen_General"] = pd.DataFrame([resumen])
        if "Validado" in df_val_data.columns:
            dfs["Validaciones_x_Validado"] = (
                df_val_data["Validado"].fillna("(blank)").astype(str).value_counts().reset_index().rename(
                    columns={"index": "Validado", "Validado": "Count"}
                )
            )

    for sheet, name in [
        ("Pivot", "Validaciones_Pivot"),
    ]:
        df_sheet = safe_read_excel(validaciones_xlsx, sheet)
        if df_sheet is not None:
            dfs[name] = df_sheet

    for sheet, name in [
        ("Pivot", "Fichas_Pivot"),
        ("Data", "Fichas_Data"),
    ]:
        df_sheet = safe_read_excel(fichas_xlsx, sheet)
        if df_sheet is not None:
            dfs[name] = df_sheet

    for sheet, name in [
        ("Pivot", "Ejecuciones_Pivot"),
        ("Data", "Ejecuciones_Data"),
    ]:
        df_sheet = safe_read_excel(ejecuciones_xlsx, sheet)
        if df_sheet is not None:
            dfs[name] = df_sheet

    for sheet, name in [
        ("Resumen", "Laura_Resumen"),
        ("Mails_Subtematica", "Laura_Subtematica"),
        ("Mails_Tematica", "Laura_Tematica"),
        ("Mails_Status", "Laura_Status"),
        ("CIFs_Segmento_Total", "Laura_Segmento_Total"),
        ("Seg_x_Subtematica", "Laura_Seg_x_Subtematica"),
        ("Seg_x_Tematica", "Laura_Seg_x_Tematica"),
    ]:
        df_sheet = safe_read_excel(laura_xlsx, sheet)
        if df_sheet is not None:
            dfs[name] = df_sheet

    for sheet, name in [
        ("Resumen", "Revision_Resumen"),
        ("Data", "Revision_Data"),
    ]:
        df_sheet = safe_read_excel(revision_xlsx, sheet)
        if df_sheet is not None:
            dfs[name] = df_sheet

    weights_path = Path(__file__).resolve().parent / "validaciones_revision_pesos.json"
    if weights_path.exists():
        try:
            weights = json.loads(weights_path.read_text(encoding="utf-8"))
            flat = []
            flat.append({"scope": "default", "key": "default", "weight": weights.get("default", 1.0)})
            for k, v in (weights.get("automatismo") or {}).items():
                flat.append({"scope": "automatismo", "key": k, "weight": v})
            for k, v in (weights.get("segmento") or {}).items():
                flat.append({"scope": "segmento", "key": k, "weight": v})
            for k, v in (weights.get("pair") or {}).items():
                flat.append({"scope": "pair", "key": k, "weight": v})
            dfs["Revision_Pesos"] = pd.DataFrame(flat)
        except Exception as exc:
            dfs["Revision_Pesos"] = pd.DataFrame([{"WARN": f"No se pudo leer pesos JSON: {exc}"}])

    with pd.ExcelWriter(out, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
        for sheet_name, df in dfs.items():
            safe_name = sheet_name[:31]
            df.to_excel(writer, sheet_name=safe_name, index=False)

    try:
        _rellenar_plantilla_stats(out, fecha, base_dir, output_dir)
    except Exception as exc:
        print(f"[WARN] No se pudo rellenar Sheet1 de la plantilla: {exc}")

    print(f"[OK] Stats generado en: {out}")


def _rellenar_plantilla_stats(stats_path: Path, fecha: str, base_dir: Path, output_dir: Path):
    laura_xlsx = output_dir / "mails_por_subtematica.xlsx"
    fichas_xlsx = output_dir / "fichas_levantadas.xlsx"

    if not laura_xlsx.exists():
        raise FileNotFoundError(f"No existe {laura_xlsx}")
    if not fichas_xlsx.exists():
        raise FileNotFoundError(f"No existe {fichas_xlsx}")

    df_seg_tematica = pd.read_excel(laura_xlsx, sheet_name="Seg_x_Tematica")
    df_seg_subtematica = pd.read_excel(laura_xlsx, sheet_name="Seg_x_Subtematica")
    df_fichas_pivot = pd.read_excel(fichas_xlsx, sheet_name="Pivot")

    def _norm(value):
        txt = str(value) if value is not None else ""
        txt = "".join(c for c in unicodedata.normalize("NFKD", txt) if not unicodedata.combining(c))
        return txt.strip().lower()

    def _find_seg_col(df: pd.DataFrame, candidates: list[str]) -> str:
        for c in candidates:
            if c in df.columns:
                return c
        raise KeyError(f"No se encontraron columnas {candidates} en {df.columns.tolist()}")

    seg_gc = _find_seg_col(df_seg_tematica, ["GGCC", "GC"])
    seg_me = _find_seg_col(df_seg_tematica, ["ME"])
    seg_pe = _find_seg_col(df_seg_tematica, ["PE"])

    seg_gc_sub = _find_seg_col(df_seg_subtematica, ["GGCC", "GC"])
    seg_me_sub = _find_seg_col(df_seg_subtematica, ["ME"])
    seg_pe_sub = _find_seg_col(df_seg_subtematica, ["PE"])

    def _build_row_map(df: pd.DataFrame, key_col: str) -> dict[str, dict]:
        out_map: dict[str, dict] = {}
        for _, r in df.iterrows():
            out_map[_norm(r.get(key_col))] = r.to_dict()
        return out_map

    map_tematica = _build_row_map(df_seg_tematica, "Location")
    map_sub = _build_row_map(df_seg_subtematica, "Sublocation")
    map_fichas = _build_row_map(df_fichas_pivot, "Automatismo")

    def _counts_from_map(m: dict[str, dict], key: str, cols: tuple[str, str, str]) -> dict[str, int]:
        row = m.get(_norm(key))
        if not row:
            return {"GC": 0, "ME": 0, "PE": 0}
        return {
            "GC": int(row.get(cols[0], 0) or 0),
            "ME": int(row.get(cols[1], 0) or 0),
            "PE": int(row.get(cols[2], 0) or 0),
        }

    def _sum_counts(m: dict[str, dict], keys: list[str], cols: tuple[str, str, str]) -> dict[str, int]:
        acc = {"GC": 0, "ME": 0, "PE": 0}
        for k in keys:
            c = _counts_from_map(m, k, cols)
            acc["GC"] += c["GC"]
            acc["ME"] += c["ME"]
            acc["PE"] += c["PE"]
        return acc

    def _agregado(counts: dict[str, int]) -> int:
        return int(counts.get("GC", 0)) + int(counts.get("ME", 0)) + int(counts.get("PE", 0))

    def _pct(n: float, d: float) -> float:
        return 0.0 if d == 0 else float(n) / float(d)

    # Total correos (fila 3) desde Seg_x_Tematica
    total_correos = {
        "GC": int(df_seg_tematica[seg_gc].sum()),
        "ME": int(df_seg_tematica[seg_me].sum()),
        "PE": int(df_seg_tematica[seg_pe].sum()),
    }

    # Cliente no encontrado (fila 4) desde ia.csv+rpa.csv (Status contiene CLIENTE_NO_ENCONTRADO)
    cliente_no_encontrado = 0
    ia_csv = base_dir / "ia.csv"
    rpa_csv = base_dir / "rpa.csv"
    if ia_csv.exists() and rpa_csv.exists():
        df_ia = pd.read_csv(ia_csv, sep=";")
        df_rpa = pd.read_csv(rpa_csv)
        merged = laura_build_merged_df(df_ia, df_rpa)
        if "Status" in merged.columns:
            cliente_no_encontrado = int(
                merged["Status"].fillna("").astype(str).str.contains("CLIENTE_NO_ENCONTRADO").sum()
            )

    # Total correo tematicas (fila 5): suma de subtematicas clave (sin no accion y sin reparacion terminales)
    tematicas = _sum_counts(
        map_sub,
        [
            "Duplicado de factura",
            "Baja Línea",
            "Alta PS",
            "Baja PS",
            "Apertura Avería Fijo",
            "Apertura Avería Móvil",
            "Desvíos Llamadas",
            "Servicio Multisim",
        ],
        (seg_gc_sub, seg_me_sub, seg_pe_sub),
    )

    # Fichas levantadas (fila 6): suma de automatismos clave
    fichas = _sum_counts(
        map_fichas,
        [
            "Duplicado de factura",
            "Baja de linea",
            "Tramitar servicios",
            "Apertura avería fijo",
            "Apertura avería móvil",
            "Desvio de llamadas",
            "Servicio multisim",
        ],
        ("GGCC", "ME", "PE"),
    )

    wb = load_workbook(stats_path)
    ws = wb["Sheet1"] if "Sheet1" in wb.sheetnames else wb.active

    # Encabezado fecha (C1)
    ws.cell(row=1, column=3).value = fecha

    def _set_row(row: int, ggcc=None, me=None, pe=None, agg=None, only_agg: bool = False):
        if not only_agg:
            ws.cell(row=row, column=3).value = ggcc
            ws.cell(row=row, column=4).value = me
            ws.cell(row=row, column=5).value = pe
        ws.cell(row=row, column=6).value = agg

    # Filas generales
    _set_row(3, total_correos["GC"], total_correos["ME"], total_correos["PE"], _agregado(total_correos))
    _set_row(4, None, None, None, cliente_no_encontrado, only_agg=True)
    _set_row(5, tematicas["GC"], tematicas["ME"], tematicas["PE"], _agregado(tematicas))
    _set_row(6, fichas["GC"], fichas["ME"], fichas["PE"], _agregado(fichas))
    _set_row(
        7,
        _pct(fichas["GC"], tematicas["GC"]),
        _pct(fichas["ME"], tematicas["ME"]),
        _pct(fichas["PE"], tematicas["PE"]),
        _pct(_agregado(fichas), _agregado(tematicas)),
    )

    def _fill_block(
        total_row: int,
        pct_row: int,
        boton_row: int,
        pct_boton_row: int,
        total_counts: dict[str, int],
        boton_counts: dict[str, int],
    ):
        total_agg = _agregado(total_counts)
        boton_agg = _agregado(boton_counts)

        _set_row(total_row, total_counts["GC"], total_counts["ME"], total_counts["PE"], total_agg)
        _set_row(
            pct_row,
            _pct(total_counts["GC"], total_correos["GC"]),
            _pct(total_counts["ME"], total_correos["ME"]),
            _pct(total_counts["PE"], total_correos["PE"]),
            _pct(total_agg, _agregado(total_correos)),
        )
        _set_row(boton_row, boton_counts["GC"], boton_counts["ME"], boton_counts["PE"], boton_agg)
        _set_row(
            pct_boton_row,
            _pct(boton_counts["GC"], total_counts["GC"]),
            _pct(boton_counts["ME"], total_counts["ME"]),
            _pct(boton_counts["PE"], total_counts["PE"]),
            _pct(boton_agg, total_agg),
        )

    # Duplicados
    dup_total = _counts_from_map(map_sub, "Duplicado de factura", (seg_gc_sub, seg_me_sub, seg_pe_sub))
    dup_boton = _counts_from_map(map_fichas, "Duplicado de factura", ("GGCC", "ME", "PE"))
    _fill_block(9, 10, 11, 12, dup_total, dup_boton)

    # Baja de linea
    baja_total = _counts_from_map(map_sub, "Baja Línea", (seg_gc_sub, seg_me_sub, seg_pe_sub))
    baja_boton = _counts_from_map(map_fichas, "Baja de linea", ("GGCC", "ME", "PE"))
    _fill_block(14, 15, 16, 17, baja_total, baja_boton)

    # AyB (Alta PS + Baja PS)
    ayb_total = _sum_counts(map_sub, ["Alta PS", "Baja PS"], (seg_gc_sub, seg_me_sub, seg_pe_sub))
    ayb_boton = _counts_from_map(map_fichas, "Tramitar servicios", ("GGCC", "ME", "PE"))
    _fill_block(19, 20, 21, 22, ayb_total, ayb_boton)

    # Averia fijo
    fijo_total = _counts_from_map(map_sub, "Apertura Avería Fijo", (seg_gc_sub, seg_me_sub, seg_pe_sub))
    fijo_boton = _counts_from_map(map_fichas, "Apertura avería fijo", ("GGCC", "ME", "PE"))
    _fill_block(24, 25, 26, 27, fijo_total, fijo_boton)

    # Averia movil
    movil_total = _counts_from_map(map_sub, "Apertura Avería Móvil", (seg_gc_sub, seg_me_sub, seg_pe_sub))
    movil_boton = _counts_from_map(map_fichas, "Apertura avería móvil", ("GGCC", "ME", "PE"))
    _fill_block(29, 30, 31, 32, movil_total, movil_boton)

    # Desvio llamadas
    desvio_total = _counts_from_map(map_sub, "Desvíos Llamadas", (seg_gc_sub, seg_me_sub, seg_pe_sub))
    desvio_boton = _counts_from_map(map_fichas, "Desvio de llamadas", ("GGCC", "ME", "PE"))
    _fill_block(34, 35, 36, 37, desvio_total, desvio_boton)

    # Multisim
    multisim_total = _counts_from_map(map_sub, "Servicio Multisim", (seg_gc_sub, seg_me_sub, seg_pe_sub))
    multisim_boton = _counts_from_map(map_fichas, "Servicio multisim", ("GGCC", "ME", "PE"))
    _fill_block(39, 40, 41, 42, multisim_total, multisim_boton)

    # Accion no requerida (tematica)
    accion_total = _counts_from_map(map_tematica, "Accion no requerida", (seg_gc, seg_me, seg_pe))
    accion_boton = _counts_from_map(map_fichas, "Accion no requerida", ("GGCC", "ME", "PE"))
    _fill_block(44, 45, 46, 47, accion_total, accion_boton)

    # Reparacion terminales
    rep_total = _counts_from_map(map_sub, "Reparacion Terminales", (seg_gc_sub, seg_me_sub, seg_pe_sub))
    rep_boton = _counts_from_map(map_fichas, "Reparacion Terminales", ("GGCC", "ME", "PE"))
    _fill_block(49, 50, 51, 52, rep_total, rep_boton)

    wb.save(stats_path)


def laura_load_inputs(folder: str, segmento_csv: str):
    ia_csv = f"{folder}/ia.csv"
    rpa_csv = f"{folder}/rpa.csv"

    df_ia = pd.read_csv(ia_csv, sep=";")
    df_segmento = pd.read_csv(segmento_csv, on_bad_lines="skip")
    df_rpa = pd.read_csv(rpa_csv)
    return df_ia, df_segmento, df_rpa


def laura_build_merged_df(df_ia: pd.DataFrame, df_rpa: pd.DataFrame) -> pd.DataFrame:
    df_rpa = df_rpa.drop_duplicates(subset=["IDU"], keep="first")
    df_rpa["IDU"] = df_rpa["IDU"].fillna("").astype(str)
    df_rpa = df_rpa.drop_duplicates(subset=["IDU"])
    df_rpa = df_rpa[df_rpa["IDU"].isin(df_ia["idLotus"])]

    merged_df = pd.merge(df_ia, df_rpa, left_on="idLotus", right_on="IDU", how="inner")
    merged_df = merged_df.drop_duplicates(subset=["IDU"])

    merged_df["Sublocation"] = merged_df.apply(
        lambda x: "Duplicado de factura" if "Duplicado de factura" in x.Sublocation else x.Sublocation, axis=1
    )
    merged_df["Location"] = merged_df.apply(
        lambda x: "Categoría no contemplada" if "Categoría no contemplada" in x.Location else x.Location, axis=1
    )
    return merged_df


def laura_compute_tables(merged_df: pd.DataFrame, df_segmento: pd.DataFrame):
    sublocation_counts = merged_df["Sublocation"].value_counts()
    location_counts = merged_df["Location"].value_counts()
    status_counts = merged_df["Status"].value_counts()

    sublocation_counts_df = sublocation_counts.reset_index().rename(
        columns={"index": "Sublocation", "Sublocation": "Count"}
    )
    location_counts_df = location_counts.reset_index().rename(columns={"index": "Location", "Location": "Count"})
    status_counts_df = status_counts.reset_index().rename(columns={"index": "Status", "Status": "Count"})

    merged_with_segmento = pd.merge(
        merged_df, df_segmento, left_on="Documento", right_on="NUMERO DOCUMENTO", how="left"
    )

    total_segment_counts = merged_with_segmento["SEGMENTO"].value_counts(dropna=False)
    total_segment_counts_df = total_segment_counts.reset_index().rename(
        columns={"index": "SEGMENTO", "SEGMENTO": "Count"}
    )

    seg_x_subtematica = (
        merged_with_segmento.groupby(["Sublocation", "SEGMENTO"]).size().unstack(fill_value=0).reset_index()
    )
    seg_x_tematica = (
        merged_with_segmento.groupby(["Location", "SEGMENTO"]).size().unstack(fill_value=0).reset_index()
    )

    return {
        "merged_df": merged_with_segmento,
        "mails_subtematica": sublocation_counts_df,
        "mails_tematica": location_counts_df,
        "mails_status": status_counts_df,
        "cifs_segmento_total": total_segment_counts_df,
        "seg_x_subtematica": seg_x_subtematica,
        "seg_x_tematica": seg_x_tematica,
    }


def laura_write_excel(output_excel: str, tables: dict):
    resumen_df = pd.DataFrame(
        [
            {"Metric": "Mails Totales (rows, cols)", "Value": str(tables["merged_df"].shape)},
            {"Metric": "Output Excel", "Value": output_excel},
        ]
    )

    with pd.ExcelWriter(output_excel, engine="openpyxl") as writer:
        resumen_df.to_excel(writer, sheet_name="Resumen", index=False)
        tables["mails_subtematica"].to_excel(writer, sheet_name="Mails_Subtematica", index=False)
        tables["mails_tematica"].to_excel(writer, sheet_name="Mails_Tematica", index=False)
        tables["mails_status"].to_excel(writer, sheet_name="Mails_Status", index=False)
        tables["cifs_segmento_total"].to_excel(writer, sheet_name="CIFs_Segmento_Total", index=False)
        tables["seg_x_subtematica"].to_excel(writer, sheet_name="Seg_x_Subtematica", index=False)
        tables["seg_x_tematica"].to_excel(writer, sheet_name="Seg_x_Tematica", index=False)


def laura_run(folder: str, segmento_csv: str, output: str | None):
    output_dir = Path(folder) / "output"
    output_dir.mkdir(parents=True, exist_ok=True)

    if output:
        output_path = Path(output)
        if not output_path.is_absolute() and str(output_path.parent) == ".":
            output_path = output_dir / output_path.name
    else:
        output_path = output_dir / "mails_por_subtematica.xlsx"

    df_ia, df_segmento, df_rpa = laura_load_inputs(folder, segmento_csv)
    merged_df = laura_build_merged_df(df_ia, df_rpa)

    print("Mails Totales:\n" + str(merged_df.shape) + "\n")
    print("Mails desglosados por subtematica:\n" + str(merged_df["Sublocation"].value_counts()))
    print("Mails desglosados por tematica:\n" + str(merged_df["Location"].value_counts()))
    print("Mails desglosados por Status:\n" + str(merged_df["Status"].value_counts()))

    tables = laura_compute_tables(merged_df, df_segmento)

    print("Total de CIFs por segmento:\n" + str(tables["merged_df"]["SEGMENTO"].value_counts(dropna=False)))
    print(
        "Conteo de CIFs de cada segmento en cada subtematica:\n"
        + str(tables["merged_df"].groupby(["Sublocation", "SEGMENTO"]).size().unstack(fill_value=0))
    )
    print(
        "Conteo de CIFs de cada segmento en cada tematica:\n"
        + str(tables["merged_df"].groupby(["Location", "SEGMENTO"]).size().unstack(fill_value=0))
    )

    laura_write_excel(str(output_path), tables)
    print(f"\nGuardado Excel con tablas en: {output_path}")


def main():
    args = parse_args()
    if args.command is None:
        args.command = "semanal"

    if args.command == "stats":
        generar_stats_desde_outputs(args.fecha, template_path=args.stats_template, output_path=args.stats_output)
        return

    if args.command == "semanal":
        fecha = args.fecha
        validados(fecha)
        fichas_levantadas(fecha)
        ejecuciones(fecha)
        if not args.no_revision:
            try:
                generar_validaciones_revision(
                    fecha, sample_size=args.revision_size, min_full_stratum=args.revision_min_full
                )
            except Exception as exc:
                print(f"[WARN] Revision validaciones fallo: {exc}")
        if not args.no_laura:
            base_dir, _ = get_dirs(fecha)
            try:
                laura_run(str(base_dir), "data/CIF-Segmento_3.csv", None)
            except FileNotFoundError as exc:
                print(f"[WARN] Laura no ejecutado: falta fichero de entrada: {exc.filename}")
            except Exception as exc:
                print(f"[WARN] Laura fallo: {exc}")
        if not args.no_stats:
            generar_stats_desde_outputs(fecha, template_path=args.stats_template, output_path=args.stats_output)
        return

    if args.command == "laura":
        laura_run(args.folder, args.segmento_csv, args.output)
        return

    raise SystemExit(f"Comando no soportado: {args.command}")


if __name__ == "__main__":
    main()

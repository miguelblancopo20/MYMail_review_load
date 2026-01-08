from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill


def aplicar_estilos_validaciones_excel(excel_path: Path) -> None:
    header_azul = {"idLotus", "Location", "Sublocation", "Subject", "Question", "MailToAgent", "Faltan datos?"}
    fill_azul = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    fill_verde = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")

    wb = load_workbook(excel_path)
    ws = wb["Data"] if "Data" in wb.sheetnames else wb.active

    for cell in ws[1]:
        if str(cell.value) in header_azul:
            cell.fill = fill_azul
        else:
            cell.fill = fill_verde

    wb.save(excel_path)


def reordenar_pivot_blank(
    pivot: pd.DataFrame, blank_label: str, after_column: str, keep_total_last: bool = False
) -> pd.DataFrame:
    cols = list(pivot.columns)
    idx = list(pivot.index)

    total_label = "Total"
    cols_wo_special = [c for c in cols if c not in (blank_label, total_label)]
    idx_wo_special = [i for i in idx if i not in (blank_label, total_label)]

    ordered_cols = cols_wo_special
    if blank_label in cols:
        if after_column in ordered_cols:
            insert_at = ordered_cols.index(after_column) + 1
        else:
            insert_at = len(ordered_cols)
        ordered_cols = ordered_cols[:insert_at] + [blank_label] + ordered_cols[insert_at:]

    if keep_total_last and total_label in cols:
        ordered_cols = [c for c in ordered_cols if c != total_label] + [total_label]

    ordered_idx = idx_wo_special
    if blank_label in idx:
        ordered_idx = ordered_idx + [blank_label]
    if keep_total_last and total_label in idx:
        ordered_idx = ordered_idx + [total_label]

    return pivot.reindex(index=ordered_idx, columns=ordered_cols)


def reordenar_pivot_blank_multiindex(
    pivot: pd.DataFrame,
    blank_label: str,
    after_segment: str,
    total_label: str = "Total",
) -> pd.DataFrame:
    if not isinstance(pivot.columns, pd.MultiIndex) or pivot.columns.nlevels != 2:
        return pivot

    ordered_idx = [i for i in pivot.index if i not in (blank_label, total_label)]
    if blank_label in pivot.index:
        ordered_idx.append(blank_label)
    if total_label in pivot.index:
        ordered_idx.append(total_label)

    lvl0 = pivot.columns.get_level_values(0).unique().tolist()
    normal_lvl0 = [v for v in lvl0 if v != total_label]

    segments = pivot.columns.get_level_values(1).unique().tolist()
    segments_normal = [s for s in segments if s not in ("", blank_label)]
    if blank_label in segments:
        insert_at = segments_normal.index(after_segment) + 1 if after_segment in segments_normal else len(segments_normal)
        segments_order = segments_normal[:insert_at] + [blank_label] + segments_normal[insert_at:]
    else:
        segments_order = segments_normal

    ordered_cols: list[tuple[str, str]] = []
    for v0 in normal_lvl0:
        for seg in segments_order:
            col = (v0, seg)
            if col in pivot.columns:
                ordered_cols.append(col)

    for v0 in normal_lvl0:
        for seg in segments_normal:
            col = (v0, seg)
            if col in pivot.columns and col not in ordered_cols:
                ordered_cols.append(col)

    if total_label in lvl0:
        total_cols = [c for c in pivot.columns if c[0] == total_label]
        ordered_cols.extend([c for c in total_cols if c not in ordered_cols])

    return pivot.reindex(index=ordered_idx, columns=ordered_cols)


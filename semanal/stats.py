from __future__ import annotations

import json
import logging
import shutil
import unicodedata
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

from .config import cfg_get, resolve_repo_path
from .paths import get_dirs, repo_root
from .subtematicas import build_merged_df

logger = logging.getLogger(__name__)


def _stats_from_validaciones(fecha: str, base_dir: Path, output_dir: Path, safe_read_excel) -> dict[str, pd.DataFrame]:
    # Paso a paso:
    # 1) Leer `validaciones.xlsx` (hoja Data) si existe.
    # 2) Generar un resumen general (totales y NO VALIDADOS).
    # 3) Generar distribucion por `Validado` (si aplica).
    # 4) Adjuntar la hoja Pivot (si existe).
    validaciones_xlsx = output_dir / str(cfg_get("outputs.validaciones_xlsx", "validaciones.xlsx"))
    dfs: dict[str, pd.DataFrame] = {}

    df_val_data = safe_read_excel(validaciones_xlsx, "Data")
    if df_val_data is not None:
        resumen = {
            "fecha": fecha,
            "folder": str(base_dir),
            "validaciones_total": int(df_val_data.shape[0]),
            "validaciones_no_validados": int((df_val_data.get("Validado") != "VALIDADO").sum())
            if "Validado" in df_val_data.columns
            else "",
        }
        dfs["Resumen_General"] = pd.DataFrame([resumen])
        if "Validado" in df_val_data.columns:
            dfs["Validaciones_x_Validado"] = (
                df_val_data["Validado"].fillna("(blank)").astype(str).value_counts().reset_index().rename(
                    columns={"index": "Validado", "Validado": "Count"}
                )
            )

    df_pivot = safe_read_excel(validaciones_xlsx, "Pivot")
    if df_pivot is not None:
        dfs["Validaciones_Pivot"] = df_pivot

    return dfs


def _build_stats_dfs(fecha: str, base_dir: Path, output_dir: Path) -> dict[str, pd.DataFrame]:
    # Paso a paso:
    # 1) Definir una lectura segura de Excels (devuelve None si no existe el fichero).
    # 2) Construir DataFrames auxiliares a partir de los outputs del semanal.
    # 3) Devolver un dict hoja->DataFrame listo para volcar a Excel.
    def safe_read_excel(path: Path, sheet: str):
        if not path.exists():
            return None
        try:
            return pd.read_excel(path, sheet_name=sheet)
        except Exception as exc:
            return pd.DataFrame([{"WARN": f"No se pudo leer {path.name}:{sheet} -> {exc}"}])

    fichas_xlsx = output_dir / str(cfg_get("outputs.fichas_levantadas_xlsx", "fichas_levantadas.xlsx"))
    ejecuciones_xlsx = output_dir / str(cfg_get("outputs.ejecuciones_xlsx", "ejecuciones.xlsx"))
    subtematicas_xlsx = output_dir / str(cfg_get("outputs.subtematicas_xlsx", "mails_por_subtematica.xlsx"))
    revision_xlsx = output_dir / str(cfg_get("outputs.revision_xlsx", "validaciones_revision.xlsx"))

    dfs: dict[str, pd.DataFrame] = {}

    # (A) Validaciones (extraido a helper)
    dfs.update(_stats_from_validaciones(fecha, base_dir, output_dir, safe_read_excel))

    # (B) Fichas / Ejecuciones (Data + Pivot)
    for sheet, name in [("Pivot", "Fichas_Pivot"), ("Data", "Fichas_Data")]:
        df_sheet = safe_read_excel(fichas_xlsx, sheet)
        if df_sheet is not None:
            dfs[name] = df_sheet

    for sheet, name in [("Pivot", "Ejecuciones_Pivot"), ("Data", "Ejecuciones_Data")]:
        df_sheet = safe_read_excel(ejecuciones_xlsx, sheet)
        if df_sheet is not None:
            dfs[name] = df_sheet

    # (C) Subtematicas (si existe)
    for sheet, name in [
        ("Resumen", "Subtematicas_Resumen"),
        ("Mails_Subtematica", "Subtematicas_Subtematica"),
        ("Mails_Tematica", "Subtematicas_Tematica"),
        ("Mails_Status", "Subtematicas_Status"),
        ("CIFs_Segmento_Total", "Subtematicas_Segmento_Total"),
        ("Seg_x_Subtematica", "Subtematicas_Seg_x_Subtematica"),
        ("Seg_x_Tematica", "Subtematicas_Seg_x_Tematica"),
    ]:
        df_sheet = safe_read_excel(subtematicas_xlsx, sheet)
        if df_sheet is not None:
            dfs[name] = df_sheet

    # (D) Revision (si existe)
    for sheet, name in [("Resumen", "Revision_Resumen"), ("Data", "Revision_Data")]:
        df_sheet = safe_read_excel(revision_xlsx, sheet)
        if df_sheet is not None:
            dfs[name] = df_sheet

    # (E) Pesos de revision (si existe)
    weights_path = repo_root() / str(cfg_get("paths.revision_weights", "validaciones_revision_pesos.json"))
    if weights_path.exists():
        try:
            weights = json.loads(weights_path.read_text(encoding="utf-8"))
            flat = [{"scope": "default", "key": "default", "weight": weights.get("default", 1.0)}]
            for k, v in (weights.get("automatismo") or {}).items():
                flat.append({"scope": "automatismo", "key": k, "weight": v})
            for k, v in (weights.get("segmento") or {}).items():
                flat.append({"scope": "segmento", "key": k, "weight": v})
            for k, v in (weights.get("pair") or {}).items():
                flat.append({"scope": "pair", "key": k, "weight": v})
            dfs["Revision_Pesos"] = pd.DataFrame(flat)
        except Exception as exc:
            dfs["Revision_Pesos"] = pd.DataFrame([{"WARN": f"No se pudo leer pesos JSON: {exc}"}])

    return dfs


def generar_stats_desde_outputs(fecha: str, template_path: str | None = None, output_path: str | None = None) -> None:
    #---6. Stats---- Copia plantilla y agrega hojas (Resumen/Pivots/Subtematicas/Revision) en `Stats_{fecha}.xlsx`.
    # Paso a paso:
    # 1) Resolver ruta de plantilla y salida.
    # 2) Copiar la plantilla al output (con fallbacks OneDrive/Excel-abierto).
    # 3) Construir hojas auxiliares desde los Excels del semanal.
    # 4) Anadir/reemplazar hojas auxiliares en el workbook de salida.
    # 5) Rellenar `Sheet1` con KPIs (si existen los Excels necesarios).
    # 6) Loggear la ruta final.
    logger.info("Stats: inicio (fecha=%s)", fecha)
    base_dir, output_dir = get_dirs(fecha)

    if template_path:
        tpl_candidates = [Path(template_path)]
    else:
        tpl_from_cfg = resolve_repo_path(cfg_get("paths.stats_template", "Plantilla_Stats.xlsx"))
        tpl_candidates = [p for p in [tpl_from_cfg, Path.cwd() / "Plantilla_Stats.xlsx", repo_root() / "Plantilla_Stats.xlsx"] if p]

    tpl = next((p for p in tpl_candidates if p.exists()), None)

    if output_path:
        out = Path(output_path)
        if not out.is_absolute():
            out = (output_dir / out.name) if str(out.parent) == "." else (Path.cwd() / out).resolve()
    else:
        pattern = str(cfg_get("outputs.stats_xlsx_pattern", "Stats_{fecha}.xlsx"))
        out = output_dir / pattern.format(fecha=fecha)

    if tpl is None:
        checked = ", ".join(str(p) for p in tpl_candidates)
        logger.warning("Stats: no existe la plantilla. Rutas comprobadas: %s", checked)
        return

    out.parent.mkdir(parents=True, exist_ok=True)

    try:
        shutil.copy2(tpl, out)
    except PermissionError as exc:
        logger.warning("Stats: no se puede leer/copiar la plantilla (esta abierta en Excel?): %s (%s)", tpl, exc)
        return
    except FileNotFoundError as exc:
        logger.warning("Stats: no se pudo copiar la plantilla (posible OneDrive/placeholder): %s (%s)", tpl, exc)
        try:
            wb_tpl = load_workbook(tpl)
            wb_tpl.save(out)
        except Exception as exc2:
            logger.warning("Stats: fallback openpyxl fallo copiando plantilla: %s", exc2)
            return
    except OSError as exc:
        logger.warning("Stats: error copiando plantilla: %s (%s)", tpl, exc)
        try:
            wb_tpl = load_workbook(tpl)
            wb_tpl.save(out)
        except Exception as exc2:
            logger.warning("Stats: fallback openpyxl fallo copiando plantilla: %s", exc2)
            return

    dfs = _build_stats_dfs(fecha, base_dir, output_dir)

    with pd.ExcelWriter(out, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
        for sheet_name, df in dfs.items():
            safe_name = sheet_name[:31]
            df.to_excel(writer, sheet_name=safe_name, index=False)

    try:
        _rellenar_plantilla_stats(out, fecha, base_dir, output_dir)
    except Exception as exc:
        logger.warning("Stats: no se pudo rellenar Sheet1 de la plantilla: %s", exc)

    logger.info("Stats: OK -> %s", out)


def _rellenar_plantilla_stats(stats_path: Path, fecha: str, base_dir: Path, output_dir: Path) -> None:
    # Paso a paso:
    # 1) Leer pivots necesarios (`Seg_x_Tematica`, `Seg_x_Subtematica`, `fichas_levantadas.xlsx`).
    # 2) Normalizar keys para buscar filas por texto (ignorando mayus/acentos).
    # 3) Calcular totales + porcentajes y rellenar celdas fijas en `Sheet1`.
    # 4) Guardar el workbook.
    subtematicas_xlsx = output_dir / str(cfg_get("outputs.subtematicas_xlsx", "mails_por_subtematica.xlsx"))
    fichas_xlsx = output_dir / str(cfg_get("outputs.fichas_levantadas_xlsx", "fichas_levantadas.xlsx"))

    if not subtematicas_xlsx.exists():
        raise FileNotFoundError(f"No existe {subtematicas_xlsx}")
    if not fichas_xlsx.exists():
        raise FileNotFoundError(f"No existe {fichas_xlsx}")

    df_seg_tematica = pd.read_excel(subtematicas_xlsx, sheet_name="Seg_x_Tematica")
    df_seg_subtematica = pd.read_excel(subtematicas_xlsx, sheet_name="Seg_x_Subtematica")
    df_fichas_pivot = pd.read_excel(fichas_xlsx, sheet_name="Pivot")

    def _norm(value):
        txt = str(value) if value is not None else ""
        txt = "".join(c for c in unicodedata.normalize("NFKD", txt) if not unicodedata.combining(c))
        return txt.strip().lower()

    def _find_seg_col(df: pd.DataFrame, candidates: list[str]) -> str:
        for c in candidates:
            if c in df.columns:
                return c
        raise KeyError(f"No se encontraron columnas {candidates} en {df.columns.tolist()}")

    seg_gc = _find_seg_col(df_seg_tematica, ["GGCC", "GC"])
    seg_me = _find_seg_col(df_seg_tematica, ["ME"])
    seg_pe = _find_seg_col(df_seg_tematica, ["PE"])

    seg_gc_sub = _find_seg_col(df_seg_subtematica, ["GGCC", "GC"])
    seg_me_sub = _find_seg_col(df_seg_subtematica, ["ME"])
    seg_pe_sub = _find_seg_col(df_seg_subtematica, ["PE"])

    def _build_row_map(df: pd.DataFrame, key_col: str) -> dict[str, dict]:
        out_map: dict[str, dict] = {}
        for _, r in df.iterrows():
            out_map[_norm(r.get(key_col))] = r.to_dict()
        return out_map

    map_tematica = _build_row_map(df_seg_tematica, "Location")
    map_sub = _build_row_map(df_seg_subtematica, "Sublocation")
    map_fichas = _build_row_map(df_fichas_pivot, "Automatismo")

    def _counts_from_map(m: dict[str, dict], key: str, cols: tuple[str, str, str]) -> dict[str, int]:
        row = m.get(_norm(key))
        if not row:
            return {"GC": 0, "ME": 0, "PE": 0}
        return {
            "GC": int(row.get(cols[0], 0) or 0),
            "ME": int(row.get(cols[1], 0) or 0),
            "PE": int(row.get(cols[2], 0) or 0),
        }

    def _sum_counts(m: dict[str, dict], keys: list[str], cols: tuple[str, str, str]) -> dict[str, int]:
        acc = {"GC": 0, "ME": 0, "PE": 0}
        for k in keys:
            c = _counts_from_map(m, k, cols)
            acc["GC"] += c["GC"]
            acc["ME"] += c["ME"]
            acc["PE"] += c["PE"]
        return acc

    def _agregado(counts: dict[str, int]) -> int:
        return int(counts.get("GC", 0)) + int(counts.get("ME", 0)) + int(counts.get("PE", 0))

    def _pct(n: float, d: float) -> float:
        return 0.0 if d == 0 else float(n) / float(d)

    total_correos = {
        "GC": int(df_seg_tematica[seg_gc].sum()),
        "ME": int(df_seg_tematica[seg_me].sum()),
        "PE": int(df_seg_tematica[seg_pe].sum()),
    }

    cliente_no_encontrado = 0
    ia_csv = base_dir / "ia.csv"
    rpa_csv = base_dir / "rpa.csv"
    if ia_csv.exists() and rpa_csv.exists():
        df_ia = pd.read_csv(ia_csv, sep=";")
        df_rpa = pd.read_csv(rpa_csv)
        merged = build_merged_df(df_ia, df_rpa)
        if "Status" in merged.columns:
            cliente_no_encontrado = int(
                merged["Status"].fillna("").astype(str).str.contains("CLIENTE_NO_ENCONTRADO").sum()
            )

    tematicas = _sum_counts(
        map_sub,
        [
            "Duplicado de factura",
            "Baja Línea",
            "Baja Línea Fija",
            "Alta PS",
            "Baja PS",
            "Apertura Avería Fijo",
            "Apertura Avería Móvil",
            "Desvíos Llamadas",
            "Servicio Multisim",
        ],
        (seg_gc_sub, seg_me_sub, seg_pe_sub),
    )

    fichas = _sum_counts(
        map_fichas,
        [
            "Duplicado de factura",
            "Baja de linea",
            "Tramitar servicios",
            "Apertura avería fijo",
            "Apertura avería móvil",
            "Desvio de llamadas",
            "Servicio multisim",
        ],
        ("GGCC", "ME", "PE"),
    )

    wb = load_workbook(stats_path)
    ws = wb["Sheet1"] if "Sheet1" in wb.sheetnames else wb.active

    ws.cell(row=1, column=3).value = fecha

    def _set_row(row: int, ggcc=None, me=None, pe=None, agg=None, only_agg: bool = False):
        if not only_agg:
            ws.cell(row=row, column=3).value = ggcc
            ws.cell(row=row, column=4).value = me
            ws.cell(row=row, column=5).value = pe
        ws.cell(row=row, column=6).value = agg

    _set_row(3, total_correos["GC"], total_correos["ME"], total_correos["PE"], _agregado(total_correos))
    _set_row(4, None, None, None, cliente_no_encontrado, only_agg=True)
    _set_row(5, tematicas["GC"], tematicas["ME"], tematicas["PE"], _agregado(tematicas))
    _set_row(6, fichas["GC"], fichas["ME"], fichas["PE"], _agregado(fichas))
    _set_row(
        7,
        _pct(fichas["GC"], tematicas["GC"]),
        _pct(fichas["ME"], tematicas["ME"]),
        _pct(fichas["PE"], tematicas["PE"]),
        _pct(_agregado(fichas), _agregado(tematicas)),
    )

    def _fill_block(
        total_row: int,
        pct_row: int,
        boton_row: int,
        pct_boton_row: int,
        total_counts: dict[str, int],
        boton_counts: dict[str, int],
    ):
        total_agg = _agregado(total_counts)
        boton_agg = _agregado(boton_counts)

        _set_row(total_row, total_counts["GC"], total_counts["ME"], total_counts["PE"], total_agg)
        _set_row(
            pct_row,
            _pct(total_counts["GC"], total_correos["GC"]),
            _pct(total_counts["ME"], total_correos["ME"]),
            _pct(total_counts["PE"], total_correos["PE"]),
            _pct(total_agg, _agregado(total_correos)),
        )
        _set_row(boton_row, boton_counts["GC"], boton_counts["ME"], boton_counts["PE"], boton_agg)
        _set_row(
            pct_boton_row,
            _pct(boton_counts["GC"], total_counts["GC"]),
            _pct(boton_counts["ME"], total_counts["ME"]),
            _pct(boton_counts["PE"], total_counts["PE"]),
            _pct(boton_agg, total_agg),
        )

    dup_total = _counts_from_map(map_sub, "Duplicado de factura", (seg_gc_sub, seg_me_sub, seg_pe_sub))
    dup_boton = _counts_from_map(map_fichas, "Duplicado de factura", ("GGCC", "ME", "PE"))
    _fill_block(9, 10, 11, 12, dup_total, dup_boton)

    baja_total = _sum_counts(map_sub, ["Baja Línea", "Baja Línea Fija"], (seg_gc_sub, seg_me_sub, seg_pe_sub))
    baja_boton = _counts_from_map(map_fichas, "Baja de linea", ("GGCC", "ME", "PE"))
    _fill_block(14, 15, 16, 17, baja_total, baja_boton)

    ayb_total = _sum_counts(map_sub, ["Alta PS", "Baja PS"], (seg_gc_sub, seg_me_sub, seg_pe_sub))
    ayb_boton = _counts_from_map(map_fichas, "Tramitar servicios", ("GGCC", "ME", "PE"))
    _fill_block(19, 20, 21, 22, ayb_total, ayb_boton)

    fijo_total = _counts_from_map(map_sub, "Apertura Avería Fijo", (seg_gc_sub, seg_me_sub, seg_pe_sub))
    fijo_boton = _counts_from_map(map_fichas, "Apertura avería fijo", ("GGCC", "ME", "PE"))
    _fill_block(24, 25, 26, 27, fijo_total, fijo_boton)

    movil_total = _counts_from_map(map_sub, "Apertura Avería Móvil", (seg_gc_sub, seg_me_sub, seg_pe_sub))
    movil_boton = _counts_from_map(map_fichas, "Apertura avería móvil", ("GGCC", "ME", "PE"))
    _fill_block(29, 30, 31, 32, movil_total, movil_boton)

    desvio_total = _counts_from_map(map_sub, "Desvíos Llamadas", (seg_gc_sub, seg_me_sub, seg_pe_sub))
    desvio_boton = _counts_from_map(map_fichas, "Desvio de llamadas", ("GGCC", "ME", "PE"))
    _fill_block(34, 35, 36, 37, desvio_total, desvio_boton)

    multisim_total = _counts_from_map(map_sub, "Servicio Multisim", (seg_gc_sub, seg_me_sub, seg_pe_sub))
    multisim_boton = _counts_from_map(map_fichas, "Servicio multisim", ("GGCC", "ME", "PE"))
    _fill_block(39, 40, 41, 42, multisim_total, multisim_boton)

    accion_total = _counts_from_map(map_tematica, "Accion no requerida", (seg_gc, seg_me, seg_pe))
    accion_boton = _counts_from_map(map_fichas, "Accion no requerida", ("GGCC", "ME", "PE"))
    _fill_block(44, 45, 46, 47, accion_total, accion_boton)

    rep_total = _counts_from_map(map_sub, "Reparacion Terminales", (seg_gc_sub, seg_me_sub, seg_pe_sub))
    rep_boton = _counts_from_map(map_fichas, "Reparacion Terminales", ("GGCC", "ME", "PE"))
    _fill_block(49, 50, 51, 52, rep_total, rep_boton)

    wb.save(stats_path)
